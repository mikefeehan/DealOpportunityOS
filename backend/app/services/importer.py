"""Real Universe importer.

Accepts a manual CSV/XLSX export (HelloData / Yardi / RealPage / analyst research),
normalizes the columns, attempts a Pima County parcel match, and upserts the records
as ``live_authorized`` data with a parcel-match confidence status.

Per the pilot's data-integrity rule, imported records are never auto-promoted to the
real call list. A Pima match lands a record in ``needs_review`` for an analyst to
confirm (-> ``verified``) or reject. Records with no parcel match land in ``no_match``.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
import zipfile
from datetime import datetime
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.app.models import Property
from backend.app.services.market_reference import market_rent_benchmark, market_rent_for
from backend.app.services.scanner import _lookup_pima_by_address
from backend.app.services.seed_data import purge_seed_data, upsert_property

try:  # chardet ships with the project; degrade gracefully if absent.
    import chardet
except Exception:  # noqa: BLE001
    chardet = None  # type: ignore[assignment]


# Canonical field -> accepted header variants (compared after normalization).
COLUMN_ALIASES: dict[str, list[str]] = {
    "external_id": ["propertyid", "yardipropertyid", "externalid", "sourceid", "apn", "parcelnumber", "parcelid"],
    "name": ["propertyname", "property", "name", "community", "communityname", "apartmentname", "buildingname", "asset", "assetname"],
    "address": ["address", "propertyaddress", "siteaddress", "streetaddress", "location", "fulladdress", "addr"],
    "city": ["propertycity", "sitecity", "city"],
    "prop_state": ["propertystate", "sitestate", "state"],
    "zip": ["zip", "zipcode", "postalcode", "propertyzip", "sitezip"],
    "submarket": ["submarket", "submkt"],
    "status": ["propertyspecialstatus", "specialstatus", "propertystatus", "status", "constructionstatus"],
    "units": ["units", "unitcount", "nounits", "noofunits", "totalunits", "numberofunits", "unit", "doors"],
    "year_built": ["yearbuilt", "built", "yrbuilt", "vintage", "yearbuiltoriginal", "constructionyear", "completiondate", "yearcompleted", "completionyear"],
    "average_rent": ["averagerent", "avgrent", "currentrent", "inplacerent", "rent", "effectiverent", "actualrent", "latestmonthlyrent", "monthlyrent", "avgmarketrent"],
    "market_rent": ["marketrent", "proformarent", "achievablerent", "askingrent", "targetrent"],
    "owner_name": ["owner", "ownername", "ownership", "ownershipentity", "trueowner"],
    "owner_city": ["ownercity", "mailingcity"],
    "owner_state": ["ownerstate", "mailingstate", "ownerst"],
    "last_sale_year": ["lastsaleyear", "yearpurchased", "saleyear", "acquired", "acquisitionyear", "purchaseyear", "latestsaledate", "saledate", "lastsaledate"],
    "source": ["source", "datasource", "provider"],
}

REQUIRED_ANY = ["address", "name"]  # need at least one identifier
DEFAULT_LAST_SALE_YEAR = 2012  # neutral placeholder when an import omits sale year

# Modeled Tucson submarket market-rent benchmarks ($/unit/mo). These are estimates
# used to derive rent upside when an export lacks a market/pro-forma rent column.
TUCSON_SUBMARKET_RENT = {
    "foothills": 1475, "oro valley": 1450, "catalina": 1425, "northwest": 1350, "marana": 1375,
    "north": 1350, "university": 1325, "downtown": 1375, "central": 1300, "northeast": 1325,
    "east": 1300, "midtown": 1275, "airport": 1100, "southwest": 1125, "south": 1100, "west": 1175,
}
DEFAULT_SUBMARKET_RENT = 1250

# Property status values that mean the asset is not yet a stabilized, operating
# acquisition target (kept out of the call list).
PREOPEN_STATUS_KEYWORDS = (
    "construction", "prospective", "planned", "proposed", "pre-leasing", "preleasing",
    "lease-up", "leaseup", "renovation", "under development", "predevelopment",
)


def _submarket_benchmark(submarket: str) -> int:
    text = (submarket or "").lower()
    for key, value in TUCSON_SUBMARKET_RENT.items():
        if key in text:
            return value
    return DEFAULT_SUBMARKET_RENT


def estimate_market_rent(submarket: str, average_rent: float, year_built: int) -> float:
    """Estimate market rent from a submarket benchmark when no market rent is given.

    Below-benchmark assets get the benchmark (upside); above-benchmark assets get a
    modest loss-to-lease, larger for older vintage. Returns 0 when in-place rent is
    unknown so no false upside is invented.
    """
    if not average_rent or average_rent <= 0:
        return 0.0
    benchmark = _submarket_benchmark(submarket)
    if benchmark > average_rent:
        return float(round(benchmark))
    uplift = 1.08 if (year_built and year_built <= 2010) else 1.04
    return float(round(average_rent * uplift))


def _is_preopen(status: str, year_built: int, current_year: int) -> bool:
    text = (status or "").lower()
    if any(keyword in text for keyword in PREOPEN_STATUS_KEYWORDS):
        return True
    return bool(year_built and year_built > current_year)


def _normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (header or "").strip().lower())


def _build_column_map(headers: list[str]) -> dict[str, str]:
    """Map canonical field -> actual header string present in the file."""
    normalized = {_normalize_header(h): h for h in headers if h}
    mapping: dict[str, str] = {}
    for canonical, variants in COLUMN_ALIASES.items():
        for variant in variants:
            if variant in normalized:
                mapping[canonical] = normalized[variant]
                break
    return mapping


def _to_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    match = re.search(r"-?\d[\d,]*", str(value))
    if not match:
        return 0
    try:
        return int(match.group(0).replace(",", ""))
    except ValueError:
        return 0


def _to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    match = re.search(r"-?\d[\d,]*\.?\d*", str(value))
    if not match:
        return 0.0
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return 0.0


def _to_year(value: Any) -> int:
    """Extract a 4-digit year from a value that may be a date, datetime, or year.

    Yardi exports dates like "2003-11-01 00:00:00"; a bare "1985" is also fine.
    """
    if value in (None, ""):
        return 0
    match = re.search(r"(19|20)\d{2}", str(value))
    return int(match.group(0)) if match else 0


def _decode_csv(content: bytes) -> str:
    if chardet is not None:
        guess = chardet.detect(content) or {}
        encoding = guess.get("encoding") or "utf-8"
    else:
        encoding = "utf-8"
    try:
        return content.decode(encoding, errors="replace")
    except (LookupError, UnicodeDecodeError):
        return content.decode("utf-8", errors="replace")


def parse_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    """Return a list of {header: value} dicts from a CSV or XLSX upload."""
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content)
    if lower.endswith(".xls"):
        raise ValueError("Legacy .xls is not supported. Re-save as .xlsx or .csv and re-upload.")
    return _parse_csv(content)


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = _decode_csv(content)
    text = text.lstrip("﻿")
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict[str, str]] = []
    for raw in reader:
        rows.append({(k or "").strip(): ("" if v is None else str(v).strip()) for k, v in raw.items()})
    return rows


def _sanitize_xlsx_styles(content: bytes) -> bytes:
    """Rewrite invalid color values in xl/styles.xml.

    Some exporters (Yardi, certain report builders) emit color rgb attributes
    that aren't 8-digit aRGB hex, which makes openpyxl refuse to open the file
    ("Colors must be aRGB hex values"). We only need cell values, so normalize
    the offending colors while preserving style indices the sheets reference.
    """
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content)) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                text = data.decode("utf-8", errors="replace")

                def _fix(match: "re.Match[str]") -> str:
                    value = match.group(1)
                    if re.fullmatch(r"[0-9A-Fa-f]{8}", value):
                        return match.group(0)
                    if re.fullmatch(r"[0-9A-Fa-f]{6}", value):
                        return f'rgb="FF{value}"'
                    return 'rgb="FF000000"'

                text = re.sub(r'rgb="([^"]*)"', _fix, text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    return out.getvalue()


def _load_xlsx_workbook(content: bytes):
    from openpyxl import load_workbook

    try:
        return load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - retry with sanitized styles before giving up
        return load_workbook(io.BytesIO(_sanitize_xlsx_styles(content)), read_only=True, data_only=True)


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    workbook = _load_xlsx_workbook(content)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [("" if cell is None else str(cell).strip()) for cell in header_row]
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        if values is None or all(cell is None or str(cell).strip() == "" for cell in values):
            continue
        record: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            cell = values[idx] if idx < len(values) else None
            record[header] = "" if cell is None else str(cell).strip()
        rows.append(record)
    workbook.close()
    return rows


def normalize_address(address: str) -> str:
    clean = re.sub(r"\s+", " ", (address or "").strip())
    clean = clean.strip(", ")
    if not clean:
        return ""
    # Ensure city/state context for Tucson parcel matching when the export omits it.
    upper = clean.upper()
    if "AZ" not in upper and "ARIZONA" not in upper:
        if "TUCSON" not in upper:
            clean = f"{clean}, Tucson, AZ"
        else:
            clean = f"{clean}, AZ"
    return clean


def _street_tokens(address: str) -> list[str]:
    head = (address or "").split(",")[0].upper()
    return [t for t in re.split(r"\s+", head) if t]


def _match_confidence(input_address: str, matched_address: str) -> float:
    a = _street_tokens(input_address)
    b = _street_tokens(matched_address)
    if not a or not b:
        return 0.0
    confidence = 0.0
    # Street number agreement is the strongest single signal.
    if a[0].isdigit() and b and a[0] == b[0]:
        confidence += 0.5
    a_rest = set(a[1:])
    b_rest = set(b[1:])
    if a_rest:
        overlap = len(a_rest & b_rest) / len(a_rest)
        confidence += 0.5 * overlap
    return round(min(confidence, 1.0), 2)


def _synthetic_parcel_id(address: str, name: str, external_id: str = "") -> str:
    # Prefer the source's own unique id so rows never collide and re-imports are
    # idempotent; fall back to address, then name.
    seed = (f"EXT:{external_id}" if external_id else normalize_address(address) or name or "unknown").upper()
    return f"IMP-{hashlib.sha1(seed.encode('utf-8')).hexdigest()[:12].upper()}"


def import_universe(
    db: Session,
    filename: str,
    content: bytes,
    source_name: str = "",
    enrich_parcels: bool = False,
    auto_verify_threshold: float = 0.9,
) -> dict[str, Any]:
    """Parse, match, and upsert an export. Returns a structured summary.

    enrich_parcels: when True, each row does a live Pima County parcel lookup
    (slow for large files). When False (default), records import fast as
    ``needs_review`` with a synthetic parcel id; matching can be run later.
    """
    rows = parse_rows(filename, content)
    if not rows:
        return {
            "status": "error",
            "error": "No rows found in the uploaded file.",
            "rows_seen": 0,
            "imported": 0,
        }

    headers = list(rows[0].keys())
    column_map = _build_column_map(headers)
    if not any(field in column_map for field in REQUIRED_ANY):
        return {
            "status": "error",
            "error": "Could not find a property name or address column. Provide at least one.",
            "rows_seen": len(rows),
            "imported": 0,
            "detected_columns": headers,
        }
    if "units" not in column_map:
        return {
            "status": "error",
            "error": "Could not find a units / unit count column. It is required.",
            "rows_seen": len(rows),
            "imported": 0,
            "detected_columns": headers,
        }

    source_label = source_name.strip() or "Manual import"
    summary = {
        "status": "ok",
        "filename": filename,
        "source_name": source_label,
        "rows_seen": len(rows),
        "imported": 0,
        "needs_review": 0,
        "verified": 0,
        "no_match": 0,
        "skipped": 0,
        "skipped_reasons": [],
        "column_map": column_map,
        "sample": [],
    }

    def get(row: dict[str, str], field: str) -> str:
        header = column_map.get(field)
        return row.get(header, "") if header else ""

    for index, row in enumerate(rows):
        name = get(row, "name").strip()
        external_id = get(row, "external_id").strip()
        # Compose a full street address from the location columns when present so
        # records outside Tucson proper (Marana, Oro Valley) match the right parcel.
        street = get(row, "address").strip()
        city = get(row, "city").strip()
        prop_state = get(row, "prop_state").strip()
        zip_code = get(row, "zip").strip()
        if street and (city or prop_state or zip_code):
            tail = " ".join(part for part in [prop_state, zip_code] if part)
            composed = ", ".join(part for part in [street, city] if part)
            address = f"{composed}, {tail}" if tail else composed
        else:
            address = normalize_address(street)
        units = _to_int(get(row, "units"))

        if not name and not address:
            summary["skipped"] += 1
            continue
        if units <= 0:
            summary["skipped"] += 1
            if len(summary["skipped_reasons"]) < 8:
                summary["skipped_reasons"].append(f"Row {index + 2}: missing/invalid units ({name or address})")
            continue

        matched_address = ""
        pima = (_lookup_pima_by_address(address) if address else None) if enrich_parcels else None
        if pima and pima.get("parcel_id"):
            confidence = 0.6  # a successful fuzzy parcel hit, pending analyst confirmation
            match_status = "verified" if confidence >= auto_verify_threshold else "needs_review"
            parcel_id = pima["parcel_id"]
        elif enrich_parcels:
            # Matching was attempted but nothing was found.
            confidence = 0.0
            match_status = "no_match"
            parcel_id = _synthetic_parcel_id(address, name, external_id)
            pima = {}
        else:
            # Fast path: imported from an authorized source, parcel not yet matched.
            confidence = 0.0
            match_status = "needs_review"
            parcel_id = _synthetic_parcel_id(address, name, external_id)
            pima = {}

        year_built = _to_year(get(row, "year_built")) or int(_to_float(pima.get("year_built", 0))) or 1985
        # No recorded sale -> treat as held since built (an original, long-hold
        # owner is a key motivation signal). Fall back to a neutral placeholder
        # only when neither a sale date nor a build year is available.
        last_sale_year = _to_year(get(row, "last_sale_year")) or year_built or DEFAULT_LAST_SALE_YEAR
        submarket = get(row, "submarket").strip() or "Tucson"
        status = get(row, "status").strip()
        average_rent = _to_float(get(row, "average_rent"))
        market_rent = _to_float(get(row, "market_rent"))
        if market_rent and not average_rent:
            average_rent = round(market_rent * 0.82, 0)
        if not market_rent:
            # Prefer a real HelloData market rent matched on address; otherwise a
            # modeled submarket benchmark. In-place (average) rent comes from the
            # export; if it's missing but we have a HelloData rent, use it as the
            # current rent so the property still shows a figure (no upside signal).
            hellodata_rent = market_rent_for(address)
            if hellodata_rent:
                market_rent = hellodata_rent
                if not average_rent:
                    average_rent = hellodata_rent
            elif average_rent and average_rent > 0:
                market_rent = estimate_market_rent(submarket, average_rent, year_built)
        assessed_value = _to_float(pima.get("assessed_value", 0)) or units * 95_000
        row_source = get(row, "source").strip()
        property_type = "Under Construction" if _is_preopen(status, year_built, datetime.utcnow().year) else "Apartments"

        payload = {
            "parcel_id": parcel_id,
            "name": name or address.split(",")[0],
            "address": address or f"{name}, Tucson, AZ",
            "units": units,
            "year_built": year_built,
            "building_sqft": int(units * 875),
            "assessed_value": assessed_value,
            "owner_name": get(row, "owner_name").strip() or pima.get("owner_name") or "Owner pending parcel match",
            "mailing_address": pima.get("mailing_address") or "Mailing address pending parcel match",
            "latitude": pima.get("latitude") or 32.2226,
            "longitude": pima.get("longitude") or -110.9747,
            "property_type": property_type,
            "submarket": submarket,
            "owner_city": get(row, "owner_city").strip() or pima.get("owner_city") or "",
            "owner_state": (get(row, "owner_state").strip() or pima.get("owner_state") or "").upper(),
            "source": f"{source_label}: {row_source}" if row_source else source_label,
            "source_name": source_label,
            "source_url": "",
            "last_sale_year": last_sale_year,
            "average_rent": average_rent,
            "market_rent": market_rent,
            "data_status": "live_authorized",
            "match_status": match_status,
            "match_confidence": confidence,
            "matched_address": matched_address if match_status != "no_match" else "",
            "last_verified_at": datetime.utcnow() if match_status == "verified" else None,
        }
        upsert_property(db, payload)

        summary["imported"] += 1
        summary[match_status if match_status in {"needs_review", "verified", "no_match"} else "needs_review"] += 1
        if len(summary["sample"]) < 8:
            summary["sample"].append(
                {
                    "name": payload["name"],
                    "address": payload["address"],
                    "units": units,
                    "match_status": match_status,
                    "parcel_id": parcel_id,
                }
            )

    db.commit()
    # Real data now exists — clear seeded demo records so they don't pollute the
    # call list or owner rollups.
    if summary["imported"] > 0:
        summary["demos_removed"] = purge_seed_data(db)
    return summary
